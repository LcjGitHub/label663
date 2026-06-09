import csv
import io
import base64
from datetime import datetime


def generate_csv_content(data):
    videos = data['videos']
    categories = data.get('categories', [])
    secondary_categories = data.get('secondary_categories', [])
    likes = data['likes']
    comments = data['comments']
    shares = data['shares']

    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_NONNUMERIC)

    writer.writerow(['视频名称', '一级分类', '二级分类', '点赞数', '评论数', '分享数', '总互动数'])

    for i in range(len(videos)):
        total = likes[i] + comments[i] + shares[i]
        primary = categories[i] if i < len(categories) else '-'
        secondary = secondary_categories[i] if i < len(secondary_categories) else '-'
        writer.writerow([
            videos[i],
            primary,
            secondary,
            likes[i],
            comments[i],
            shares[i],
            total
        ])

    total_likes = sum(likes)
    total_comments = sum(comments)
    total_shares = sum(shares)
    grand_total = total_likes + total_comments + total_shares

    writer.writerow(['合计', '', '', total_likes, total_comments, total_shares, grand_total])

    return output.getvalue()


def create_download_link(content, filename_prefix='互动数据', file_type='csv'):
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if file_type == 'csv':
        filename = f'{filename_prefix}_{timestamp}.csv'
        content_bytes = content.encode('utf-8-sig')
        mime_type = 'text/csv'
    else:
        filename = f'{filename_prefix}_{timestamp}.xlsx'
        content_bytes = content
        mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    b64 = base64.b64encode(content_bytes).decode()

    return dict(
        content=b64,
        filename=filename,
        type=mime_type,
        base64=True
    )


def export_data_to_csv(data, period_label='今日'):
    csv_content = generate_csv_content(data)
    return create_download_link(csv_content, f'互动数据_{period_label}', file_type='csv')


def generate_excel_content(data, category_summary_data=None, trend_data=None, category_bar_data=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    total_font = Font(bold=True, size=12)
    total_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')

    ws_raw = wb.active
    ws_raw.title = '原始数据'

    videos = data['videos']
    categories = data.get('categories', [])
    secondary_categories = data.get('secondary_categories', [])
    likes = data['likes']
    comments = data['comments']
    shares = data['shares']

    raw_headers = ['视频名称', '一级分类', '二级分类', '点赞数', '评论数', '分享数', '总互动数']
    for col_idx, header in enumerate(raw_headers, 1):
        cell = ws_raw.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i in range(len(videos)):
        total = likes[i] + comments[i] + shares[i]
        primary = categories[i] if i < len(categories) else '-'
        secondary = secondary_categories[i] if i < len(secondary_categories) else '-'
        row_data = [videos[i], primary, secondary, likes[i], comments[i], shares[i], total]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_raw.cell(row=i + 2, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border = thin_border

    total_likes = sum(likes)
    total_comments = sum(comments)
    total_shares = sum(shares)
    grand_total = total_likes + total_comments + total_shares

    total_row_idx = len(videos) + 2
    total_row_data = ['合计', '', '', total_likes, total_comments, total_shares, grand_total]
    for col_idx, value in enumerate(total_row_data, 1):
        cell = ws_raw.cell(row=total_row_idx, column=col_idx, value=value)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center_align
        cell.border = thin_border

    for col_idx in range(1, len(raw_headers) + 1):
        ws_raw.column_dimensions[get_column_letter(col_idx)].width = 18

    ws_summary = wb.create_sheet('统计摘要')

    summary_headers = ['指标', '数值']
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    summary_data = [
        ['视频总数', len(videos)],
        ['总点赞数', total_likes],
        ['总评论数', total_comments],
        ['总分享数', total_shares],
        ['总互动数', grand_total],
        ['平均点赞数', round(total_likes / len(videos), 2) if videos else 0],
        ['平均评论数', round(total_comments / len(videos), 2) if videos else 0],
        ['平均分享数', round(total_shares / len(videos), 2) if videos else 0],
        ['平均互动数', round(grand_total / len(videos), 2) if videos else 0]
    ]

    for row_idx, (metric, value) in enumerate(summary_data, 2):
        cell1 = ws_summary.cell(row=row_idx, column=1, value=metric)
        cell1.alignment = center_align
        cell1.border = thin_border
        cell2 = ws_summary.cell(row=row_idx, column=2, value=value)
        cell2.alignment = center_align
        cell2.border = thin_border

    if category_summary_data:
        current_row = len(summary_data) + 4
        ws_summary.cell(row=current_row, column=1, value='分类统计').font = Font(bold=True, size=14)
        current_row += 1

        cat_headers = ['视频分类', '视频数量', '平均点赞', '平均评论', '平均分享']
        for col_idx, header in enumerate(cat_headers, 1):
            cell = ws_summary.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        current_row += 1

        for cat_row in category_summary_data:
            row_data = [
                cat_row.get('category', ''),
                int(cat_row.get('video_count', '0').split(' ')[0]) if isinstance(cat_row.get('video_count', ''), str) else cat_row.get('video_count', 0),
                float(str(cat_row.get('avg_likes', 0)).replace(',', '')) if cat_row.get('avg_likes') else 0,
                float(str(cat_row.get('avg_comments', 0)).replace(',', '')) if cat_row.get('avg_comments') else 0,
                float(str(cat_row.get('avg_shares', 0)).replace(',', '')) if cat_row.get('avg_shares') else 0
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=current_row, column=col_idx, value=value)
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 18
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 15

    ws_chart = wb.create_sheet('图表数据')

    chart_headers = ['视频名称', '点赞数', '评论数', '分享数', '总互动数']
    for col_idx, header in enumerate(chart_headers, 1):
        cell = ws_chart.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i in range(len(videos)):
        total = likes[i] + comments[i] + shares[i]
        row_data = [videos[i], likes[i], comments[i], shares[i], total]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_chart.cell(row=i + 2, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border = thin_border

    current_row = len(videos) + 4
    ws_chart.cell(row=current_row, column=1, value='饼图数据（各视频互动占比）').font = Font(bold=True, size=14)
    current_row += 1

    pie_headers = ['视频名称', '总互动数', '占比(%)']
    for col_idx, header in enumerate(pie_headers, 1):
        cell = ws_chart.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    current_row += 1

    for i in range(len(videos)):
        total = likes[i] + comments[i] + shares[i]
        percentage = round(total / grand_total * 100, 2) if grand_total > 0 else 0
        row_data = [videos[i], total, percentage]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_chart.cell(row=current_row, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border = thin_border
        current_row += 1

    if category_bar_data:
        current_row += 2
        ws_chart.cell(row=current_row, column=1, value='分类对比柱状图数据').font = Font(bold=True, size=14)
        current_row += 1

        bar_headers = ['视频分类', '平均点赞', '平均评论', '平均分享']
        for col_idx, header in enumerate(bar_headers, 1):
            cell = ws_chart.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        current_row += 1

        cat_names, cat_avg_likes, cat_avg_comments, cat_avg_shares = category_bar_data
        for i in range(len(cat_names)):
            row_data = [cat_names[i], cat_avg_likes[i], cat_avg_comments[i], cat_avg_shares[i]]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_chart.cell(row=current_row, column=col_idx, value=value)
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1

    if trend_data:
        current_row += 2
        ws_chart.cell(row=current_row, column=1, value='趋势数据').font = Font(bold=True, size=14)
        current_row += 1

        trend_headers = ['时间段', '点赞合计', '评论合计', '分享合计', '总互动合计']
        for col_idx, header in enumerate(trend_headers, 1):
            cell = ws_chart.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        current_row += 1

        periods = trend_data.get('periods', [])
        trend_likes = trend_data.get('total_likes', [])
        trend_comments = trend_data.get('total_comments', [])
        trend_shares = trend_data.get('total_shares', [])

        for i in range(len(periods)):
            t_total = trend_likes[i] + trend_comments[i] + trend_shares[i]
            row_data = [periods[i], trend_likes[i], trend_comments[i], trend_shares[i], t_total]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_chart.cell(row=current_row, column=col_idx, value=value)
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1

    for col_idx in range(1, 6):
        ws_chart.column_dimensions[get_column_letter(col_idx)].width = 20

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_data_to_excel(data, period_label='今日', category_summary_data=None, trend_data=None, category_bar_data=None):
    excel_content = generate_excel_content(data, category_summary_data, trend_data, category_bar_data)
    return create_download_link(excel_content, f'互动数据_{period_label}', file_type='excel')
